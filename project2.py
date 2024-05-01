import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

file = openpyxl.load_workbook('C:/Users/Vyacheslav/Desktop/testSubjects.xlsx')
currentSheet = file.active

searchName = input("Введите ФИО для поиска: ")
#for row in currentSheet.iter_rows(values_only=True):
#    for value in row:
#        if value == searchName:
#            print(currentSheet.cell(row=row, column=5).value)
for i in range(1, 14):
    if currentSheet.cell(row=i, column=2).value == searchName:
        print(currentSheet.cell(row=i, column=2).value)
        j = 1
        while (j==1):
            print("Предмет: ",currentSheet.cell(row=i, column=5).value, " Дата сдачи: ",currentSheet.cell(row=i, column=6).value)
            i += 1
            if currentSheet.cell(row=i, column=2).value == None:
                continue
            else: break


